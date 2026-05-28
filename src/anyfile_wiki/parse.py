from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
import hashlib
import json
import re
import subprocess
import sys

from .scan import ScanEntry


TEXT_EXTENSIONS = {
    ".txt",
    ".md",
    ".rst",
    ".log",
    ".csv",
    ".json",
    ".yaml",
    ".yml",
    ".html",
    ".htm",
    ".py",
    ".js",
    ".ts",
    ".tsx",
    ".jsx",
    ".css",
    ".scss",
    ".go",
    ".rs",
    ".java",
    ".c",
    ".cpp",
    ".h",
    ".hpp",
    ".ps1",
    ".sh",
    ".sql",
    ".toml",
    ".xml",
}

OCR_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp"}
SPREADSHEET_EXTENSIONS = {".xls", ".xlsx"}


@dataclass(frozen=True)
class ParseJob:
    path: Path
    parser: str
    reason: str
    embedding_allowed: bool
    source_policy: str = "allow"
    source_size_bytes: int | None = None
    source_mtime: float | None = None


@dataclass(frozen=True)
class ExtractResult:
    path: str
    parser: str
    status: str
    output_path: str | None
    error: str | None
    embedding_allowed: bool
    created_at: str
    source_size_bytes: int | None = None
    source_mtime: float | None = None
    output_sha256: str | None = None
    skip_reason: str | None = None


@dataclass(frozen=True)
class ExtractPlan:
    jobs: list[ParseJob]
    skipped: list[ExtractResult]


def build_parse_jobs(entries: list[ScanEntry]) -> list[ParseJob]:
    """Build parser jobs only for entries whose policy allows content reads."""

    jobs: list[ParseJob] = []
    for entry in entries:
        decision = entry.decision
        if entry.is_dir or not decision.is_read_allowed or not decision.is_extract_allowed:
            continue
        parser = choose_parser(entry.extension)
        if parser is None:
            continue
        jobs.append(
            ParseJob(
                path=Path(entry.path),
                parser=parser,
                reason=f"{decision.access_policy}: {decision.reason}",
                embedding_allowed=decision.is_embedding_allowed,
                source_policy=decision.access_policy,
                source_size_bytes=entry.size_bytes,
                source_mtime=entry.mtime,
            )
        )
    return jobs


def build_parse_jobs_from_records(records: list[dict]) -> list[ParseJob]:
    return plan_parse_jobs_from_records(records).jobs


def plan_parse_jobs_from_records(
    records: list[dict],
    *,
    latest_success_by_path: dict[str, dict] | None = None,
    latest_by_path: dict[str, dict] | None = None,
    force: bool = False,
    retry_failed: bool = False,
) -> ExtractPlan:
    latest_success_by_path = latest_success_by_path or {}
    latest_by_path = latest_by_path or {}
    jobs: list[ParseJob] = []
    skipped: list[ExtractResult] = []
    for record in records:
        if record.get("is_dir"):
            continue
        if not record.get("is_read_allowed") or not record.get("is_extract_allowed"):
            continue
        parser = choose_parser(str(record.get("extension", "")))
        if parser is None:
            continue
        job = ParseJob(
            path=Path(str(record["path"])),
            parser=parser,
            reason=f"{record.get('access_policy')}: {record.get('policy_reason', '')}",
            embedding_allowed=bool(record.get("is_embedding_allowed")),
            source_policy=str(record.get("access_policy", "allow")),
            source_size_bytes=_optional_int(record.get("size_bytes")),
            source_mtime=_optional_float(record.get("mtime")),
        )
        latest = latest_by_path.get(str(job.path))
        latest_success = latest_success_by_path.get(str(job.path))

        if retry_failed:
            if latest and latest.get("status") in {"error", "skipped"}:
                jobs.append(job)
            else:
                skipped.append(_skip_result(job, latest_success, "retry_failed filter"))
            continue

        if force:
            jobs.append(job)
            continue

        if latest_success and _same_source(job, latest_success):
            skipped.append(_skip_result(job, latest_success, "source unchanged"))
            continue

        jobs.append(job)
    return ExtractPlan(jobs=jobs, skipped=skipped)


def choose_parser(extension: str) -> str | None:
    normalized = extension.lower()
    if normalized in TEXT_EXTENSIONS:
        return "direct_text"
    if normalized in SPREADSHEET_EXTENSIONS:
        return "spreadsheet"
    if normalized in {".pdf", ".docx", ".pptx"}:
        return "markitdown"
    if normalized in OCR_IMAGE_EXTENSIONS:
        return "ocr"
    return None


def extract_jobs(
    jobs: list[ParseJob],
    output_dir: str | Path,
    *,
    timeout_seconds: int | None = None,
) -> list[ExtractResult]:
    root = Path(output_dir)
    root.mkdir(parents=True, exist_ok=True)
    return [extract_job(job, root, timeout_seconds=timeout_seconds) for job in jobs]


def extract_job(job: ParseJob, output_dir: str | Path, *, timeout_seconds: int | None = None) -> ExtractResult:
    root = Path(output_dir)
    root.mkdir(parents=True, exist_ok=True)
    if job.parser == "direct_text":
        return _extract_direct_text(job, root)
    if job.parser == "markitdown":
        if timeout_seconds and timeout_seconds > 0:
            return _extract_with_timeout(job, root, timeout_seconds)
        return _extract_markitdown(job, root)
    if job.parser == "spreadsheet":
        if timeout_seconds and timeout_seconds > 0:
            return _extract_with_timeout(job, root, timeout_seconds)
        return _extract_spreadsheet(job, root)
    if job.parser == "ocr":
        if timeout_seconds and timeout_seconds > 0:
            return _extract_with_timeout(job, root, timeout_seconds)
        return _extract_ocr(job, root)
    return _result(job, "skipped", None, f"unsupported parser: {job.parser}")


def write_manifest(results: list[ExtractResult], path: str | Path) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8") as handle:
        for result in results:
            handle.write(json.dumps(asdict(result), ensure_ascii=False, sort_keys=True) + "\n")


def _extract_direct_text(job: ParseJob, output_dir: Path) -> ExtractResult:
    try:
        text = _read_text(job.path)
        output = output_dir / "text" / _artifact_name(job.path)
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(text, encoding="utf-8")
        return _result(job, "ok", output, None)
    except Exception as exc:  # noqa: BLE001 - manifest should capture parser failures.
        return _result(job, "error", None, str(exc))


def _extract_markitdown(job: ParseJob, output_dir: Path) -> ExtractResult:
    try:
        from markitdown import MarkItDown  # type: ignore
    except Exception as exc:  # noqa: BLE001 - optional dependency.
        return _result(job, "skipped", None, f"markitdown unavailable: {exc}")

    try:
        converter = MarkItDown()
        converted = converter.convert(str(job.path))
        text = getattr(converted, "text_content", None)
        if text is None:
            text = str(converted)
        output = output_dir / "markitdown" / _artifact_name(job.path)
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(text, encoding="utf-8")
        return _result(job, "ok", output, None)
    except Exception as exc:  # noqa: BLE001 - manifest should capture parser failures.
        return _result(job, "error", None, str(exc))


def _extract_spreadsheet(job: ParseJob, output_dir: Path) -> ExtractResult:
    try:
        if job.path.suffix.lower() == ".xlsx":
            text = _extract_xlsx_preview(job.path)
        else:
            text = _extract_xls_preview(job.path)
        output = output_dir / "spreadsheet" / _artifact_name(job.path)
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(text, encoding="utf-8")
        return _result(job, "ok", output, None)
    except ImportError as exc:
        return _result(job, "skipped", None, f"spreadsheet parser unavailable: {exc}")
    except Exception as exc:  # noqa: BLE001 - manifest should capture parser failures.
        return _result(job, "error", None, str(exc))


def _extract_ocr(job: ParseJob, output_dir: Path) -> ExtractResult:
    try:
        from rapidocr import RapidOCR  # type: ignore
    except Exception as exc:  # noqa: BLE001 - optional dependency.
        return _result(job, "skipped", None, f"rapidocr unavailable: {exc}")

    try:
        engine = RapidOCR()
        recognized = engine(str(job.path))
        lines = [str(item).strip() for item in (getattr(recognized, "txts", None) or []) if str(item).strip()]
        scores = [float(item) for item in (getattr(recognized, "scores", None) or [])]
        if not lines:
            return _result(job, "skipped", None, "ocr produced no text")

        average_score = sum(scores) / len(scores) if scores else None
        output = output_dir / "ocr" / _artifact_name(job.path)
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(_format_ocr_markdown(job, lines, average_score), encoding="utf-8")
        return _result(job, "ok", output, None)
    except Exception as exc:  # noqa: BLE001 - manifest should capture parser failures.
        return _result(job, "error", None, str(exc))


def _extract_with_timeout(job: ParseJob, output_dir: Path, timeout_seconds: int) -> ExtractResult:
    command = [
        sys.executable,
        "-m",
        "anyfile_wiki.extract_worker",
        "--path",
        str(job.path),
        "--parser",
        job.parser,
        "--out",
        str(output_dir),
        "--embedding-allowed",
        "1" if job.embedding_allowed else "0",
        "--source-policy",
        job.source_policy,
        "--source-size-bytes",
        "" if job.source_size_bytes is None else str(job.source_size_bytes),
        "--source-mtime",
        "" if job.source_mtime is None else str(job.source_mtime),
    ]
    try:
        completed = subprocess.run(
            command,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=timeout_seconds,
            check=False,
        )
    except subprocess.TimeoutExpired:
        return _result(job, "error", None, f"{job.parser} timed out after {timeout_seconds}s")

    if completed.returncode != 0:
        detail = (completed.stderr or completed.stdout or "").strip()
        return _result(job, "error", None, f"{job.parser} worker failed: {detail[:4000]}")

    try:
        payload = json.loads(completed.stdout)
        return ExtractResult(**payload)
    except Exception as exc:  # noqa: BLE001 - malformed worker output should be visible.
        detail = (completed.stdout or completed.stderr or "").strip()
        return _result(job, "error", None, f"{job.parser} worker returned invalid output: {exc}; {detail[:2000]}")


def _format_ocr_markdown(job: ParseJob, lines: list[str], average_score: float | None) -> str:
    score_line = "" if average_score is None else f"- average_score: {average_score:.4f}\n"
    text = "\n".join(lines)
    return (
        f"# OCR: {job.path.name}\n\n"
        f"- source: `{job.path}`\n"
        f"- parser: rapidocr\n"
        f"- lines: {len(lines)}\n"
        f"{score_line}\n"
        "## Text\n\n"
        f"{text}\n"
    )


def _extract_xlsx_preview(path: Path) -> str:
    from openpyxl import load_workbook  # type: ignore

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        lines = _spreadsheet_header(path, len(workbook.sheetnames), "openpyxl")
        for sheet_name in workbook.sheetnames[:20]:
            sheet = workbook[sheet_name]
            rows = _sheet_preview_rows(sheet.iter_rows(max_row=30, max_col=12, values_only=True), limit=12)
            lines.extend(_format_sheet_preview(sheet_name, sheet.max_row, sheet.max_column, rows))
        if len(workbook.sheetnames) > 20:
            lines.append(f"\n_还有 {len(workbook.sheetnames) - 20} 个工作表未展开。_")
        return "\n".join(lines).rstrip() + "\n"
    finally:
        close = getattr(workbook, "close", None)
        if callable(close):
            close()


def _extract_xls_preview(path: Path) -> str:
    import xlrd  # type: ignore

    workbook = xlrd.open_workbook(str(path), on_demand=True)
    try:
        sheet_names = workbook.sheet_names()
        lines = _spreadsheet_header(path, len(sheet_names), "xlrd")
        for sheet_name in sheet_names[:20]:
            sheet = workbook.sheet_by_name(sheet_name)
            raw_rows = (
                [sheet.cell_value(row_index, col_index) for col_index in range(min(sheet.ncols, 12))]
                for row_index in range(min(sheet.nrows, 30))
            )
            rows = _sheet_preview_rows(raw_rows, limit=12)
            lines.extend(_format_sheet_preview(sheet_name, sheet.nrows, sheet.ncols, rows))
        if len(sheet_names) > 20:
            lines.append(f"\n_还有 {len(sheet_names) - 20} 个工作表未展开。_")
        return "\n".join(lines).rstrip() + "\n"
    finally:
        release = getattr(workbook, "release_resources", None)
        if callable(release):
            release()


def _spreadsheet_header(path: Path, sheet_count: int, engine: str) -> list[str]:
    return [
        f"# 表格摘要: {path.name}",
        "",
        f"- source: `{path}`",
        f"- parser: spreadsheet_preview",
        f"- engine: {engine}",
        f"- sheets: {sheet_count}",
        "",
    ]


def _sheet_preview_rows(rows, *, limit: int) -> list[list[str]]:
    preview: list[list[str]] = []
    for row in rows:
        values = [_cell_text(value) for value in row]
        if any(values):
            preview.append(values)
        if len(preview) >= limit:
            break
    return preview


def _format_sheet_preview(sheet_name: str, row_count: int | None, col_count: int | None, rows: list[list[str]]) -> list[str]:
    lines = [
        f"## 工作表: {sheet_name}",
        "",
        f"- rows: {row_count or 0}",
        f"- columns: {col_count or 0}",
        "",
    ]
    if not rows:
        lines.extend(["_未发现可预览的非空单元格。_", ""])
        return lines

    width = max(len(row) for row in rows)
    padded = [row + [""] * (width - len(row)) for row in rows]
    header = padded[0]
    body = padded[1:] or [[""] * width]
    lines.append("| " + " | ".join(_escape_table_cell(cell or f"列{index + 1}") for index, cell in enumerate(header)) + " |")
    lines.append("| " + " | ".join("---" for _ in range(width)) + " |")
    for row in body[:11]:
        lines.append("| " + " | ".join(_escape_table_cell(cell) for cell in row) + " |")
    lines.append("")
    return lines


def _cell_text(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", " ").replace("\n", " ").replace("\r", " ").strip()
    return _truncate_cell(text, 120)


def _escape_table_cell(text: str) -> str:
    return text.replace("|", "\\|")


def _truncate_cell(text: str, limit: int) -> str:
    return text if len(text) <= limit else text[: limit - 1].rstrip() + "…"


def _result(job: ParseJob, status: str, output_path: Path | None, error: str | None) -> ExtractResult:
    return ExtractResult(
        path=str(job.path),
        parser=job.parser,
        status=status,
        output_path=str(output_path) if output_path else None,
        error=error,
        embedding_allowed=job.embedding_allowed,
        created_at=datetime.now(timezone.utc).isoformat(),
        source_size_bytes=job.source_size_bytes,
        source_mtime=job.source_mtime,
        output_sha256=_sha256_file(output_path) if output_path and output_path.exists() else None,
    )


def _skip_result(job: ParseJob, latest_success: dict | None, reason: str) -> ExtractResult:
    return ExtractResult(
        path=str(job.path),
        parser=job.parser,
        status="up_to_date",
        output_path=str(latest_success.get("output_path")) if latest_success and latest_success.get("output_path") else None,
        error=None,
        embedding_allowed=job.embedding_allowed,
        created_at=datetime.now(timezone.utc).isoformat(),
        source_size_bytes=job.source_size_bytes,
        source_mtime=job.source_mtime,
        output_sha256=str(latest_success.get("output_sha256")) if latest_success and latest_success.get("output_sha256") else None,
        skip_reason=reason,
    )


def _read_text(path: Path) -> str:
    data = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            return _normalize_newlines(data.decode(encoding))
        except UnicodeDecodeError:
            continue
    return _normalize_newlines(data.decode("utf-8", errors="replace"))


def _normalize_newlines(text: str) -> str:
    return text.replace("\r\n", "\n").replace("\r", "\n")


def _artifact_name(path: Path) -> str:
    digest = hashlib.sha256(str(path.resolve()).encode("utf-8", errors="ignore")).hexdigest()[:16]
    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", path.stem).strip("._") or "file"
    return f"{digest}-{stem}.md"


def _same_source(job: ParseJob, latest: dict) -> bool:
    return (
        latest.get("parser") == job.parser
        and _optional_int(latest.get("source_size_bytes")) == job.source_size_bytes
        and _optional_float(latest.get("source_mtime")) == job.source_mtime
        and bool(latest.get("output_path"))
    )


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _optional_int(value) -> int | None:
    if value is None:
        return None
    return int(value)


def _optional_float(value) -> float | None:
    if value is None:
        return None
    return float(value)
